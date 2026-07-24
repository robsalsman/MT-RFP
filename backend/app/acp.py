"""ACP wind-down need signal. The Affordable Connectivity Program paid for
home internet for 23M households until it ended in mid-2024. USAC's final
enrollment-by-zip snapshot tells us exactly how many households in a
district's zip lost that subsidy — a concrete, citable local need for
hotspot lending ("4,200 households in your area lost their internet
subsidy").
"""
import io
import logging

import httpx

from . import db

log = logging.getLogger(__name__)

# USAC's last full enrollment snapshot before the program froze
ACP_XLSX = ("https://www.usac.org/wp-content/uploads/about/documents/acp/"
            "ACP-Enrollments-by-Zip-as-of-February-8-2024.xlsx")

_SCHEMA = """CREATE TABLE IF NOT EXISTS acp_zip (
    zip TEXT PRIMARY KEY,
    households INTEGER
);"""


def ensure_loaded() -> int:
    """Import the snapshot once; returns row count in the table."""
    with db.closing_conn() as conn:
        conn.execute(_SCHEMA)
        conn.commit()
        n = conn.execute("SELECT COUNT(*) FROM acp_zip").fetchone()[0]
    if n > 1000:
        return n
    try:
        import openpyxl
        r = httpx.get(ACP_XLSX, headers={"User-Agent": "Mozilla/5.0"},
                      timeout=180, follow_redirects=True)
        r.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        log.warning("ACP snapshot download failed: %s", e)
        return n
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        try:
            z = int(row[1]); hh = int(row[2] or 0)
        except (TypeError, ValueError, IndexError):
            continue
        if z <= 0 or hh <= 0:
            continue
        rows.append((f"{z:05d}", hh))
    with db.closing_conn() as conn:
        conn.executemany(
            "INSERT OR REPLACE INTO acp_zip (zip, households) VALUES (?,?)",
            rows)
        conn.commit()
        n = conn.execute("SELECT COUNT(*) FROM acp_zip").fetchone()[0]
    log.info("ACP snapshot loaded: %d zips", n)
    return n


def households_for_zip(zip5: str | None) -> int | None:
    """Households that lost ACP in this zip (None when unknown/redacted)."""
    if not zip5:
        return None
    ensure_loaded()
    with db.closing_conn() as conn:
        row = conn.execute("SELECT households FROM acp_zip WHERE zip=?",
                           (str(zip5)[:5],)).fetchone()
    return row[0] if row else None


def impact(state: str | None = None, zips: list[str] | None = None,
           zip_prefixes: list[str] | None = None, limit: int = 15) -> dict:
    """ACP-loss picture for an area — the need map for hotspot lending."""
    ensure_loaded()
    out: list[dict] = []
    with db.closing_conn() as conn:
        if zips:
            for z in zips[:50]:
                row = conn.execute(
                    "SELECT households FROM acp_zip WHERE zip=?",
                    (str(z)[:5].zfill(5),)).fetchone()
                out.append({"zip": str(z)[:5].zfill(5),
                            "households_lost_acp": row[0] if row else None})
        elif zip_prefixes:
            for p in zip_prefixes[:20]:
                rows = conn.execute(
                    "SELECT zip, households FROM acp_zip WHERE zip LIKE ? "
                    "ORDER BY households DESC LIMIT ?",
                    (f"{p}%", limit)).fetchall()
                out += [{"zip": r[0], "households_lost_acp": r[1]}
                        for r in rows]
        elif state:
            # no state column in the snapshot: join through the board's zips
            rows = conn.execute(
                """SELECT DISTINCT c.zip FROM competitor_leads c
                   WHERE c.state=? AND c.zip IS NOT NULL""",
                (state.upper(),)).fetchall()
            for (z,) in rows[:200]:
                hit = conn.execute(
                    "SELECT households FROM acp_zip WHERE zip=?",
                    (z,)).fetchone()
                if hit:
                    out.append({"zip": z, "households_lost_acp": hit[0]})
            out.sort(key=lambda x: -(x["households_lost_acp"] or 0))
            out = out[:limit]
    total = sum(x["households_lost_acp"] or 0 for x in out)
    return {"note": "Households enrolled in ACP when the program ended "
                    "(Feb 2024 snapshot) — families who lost their home "
                    "internet subsidy. Cite this as local need for hotspot "
                    "lending.",
            "total_households": total, "zips": out}
